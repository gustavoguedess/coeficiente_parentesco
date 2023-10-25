#%% Abrir arquivo de entrada e ajuste de dados
import pandas as pd

EXCEL_FILE_NAME = 'coelhos.xlsx'

df = pd.read_excel(EXCEL_FILE_NAME, dtype=str)
df = df.rename(columns={
    'Indivíduo': 'individuo',
    'Pai': 'pai',
    'Mae': 'mae',
    'SEXO': 'sexo',
})
df = df[['individuo', 'pai', 'mae', 'sexo']]

df.replace('x', 'X', inplace=True)
df.replace('X', '', inplace=True)
df['sexo'] = df['sexo'].str.upper()

df['individuo'] = df['individuo'].apply(lambda x: x.rjust(10, ' ') if x.isnumeric() else x)
df.sort_values(by=['individuo'], inplace=True)
df['individuo'] = df['individuo'].str.strip()

df['individuo'].tolist()

# %% Classe Coelho
from collections import defaultdict

class Coelho:
    def __init__(self, nome, sexo):
        self.nome = nome
        self.sexo = sexo
        self.pai = None
        self.mae = None
        self.filhos = []

    def add_pai(self, pai):
        if pai.get_sexo() != 'M':
            raise Exception(f'Pai do coelho {self.nome}, coelho {pai.nome}, é fêmea')
        self.pai = pai
        pai.add_filho(self)
    
    def add_mae(self, mae):
        if mae.get_sexo() != 'F':
            raise Exception(f'Mãe do coelho {self.nome}, coelho {mae.nome}, é macho')
        self.mae = mae
        mae.add_filho(self)

    def add_filho(self, filho):
        self.filhos.append(filho)

    def get_arvore(self, nivel=0):
        arvore = f'{"  "*nivel}{self.nome} ({self.sexo})\n'
        for filho in self.filhos:
            arvore += filho.get_arvore(nivel+1)
        return arvore
    def get_pai(self):
        return self.pai
    def get_mae(self):
        return self.mae
    def get_sexo(self):
        return self.sexo
    
# %% Criação da árvore genealógica
coelhos = {}
for index, row in df.iterrows():
    nome = row['individuo']
    sexo = row['sexo']
    coelhos[nome] = Coelho(nome, sexo)

for index, row in df.iterrows():
    nome = row['individuo']
    pai = row['pai']
    mae = row['mae']
    if pai:
        coelhos[nome].add_pai(coelhos[pai])
    if mae:
        coelhos[nome].add_mae(coelhos[mae])

for coelho in coelhos.values():
    if coelho.get_pai() == None and coelho.get_mae() == None:
        print(coelho.get_arvore())

#%% Classe CoeficienteParentesco
class CoeficienteParentesco:
    def __init__(self, coelhos):
        self.coelhos = coelhos
        self.coeficientes_parentesco = []
    
    def calc_coeficiente_macho(self, coelho_inicial, coelho_atual=None, grau=0):
        if coelho_atual == None: coelho_atual = coelho_inicial
        if self.visitados[coelho_atual.nome]:
            return
        self.visitados[coelho_atual.nome] = True

        # Se o coelho atual for do sexo oposto ao coelho inicial
        if coelho_atual.get_sexo() != coelho_inicial.get_sexo():
            self.coeficiente_parentesco[coelho_inicial.nome][coelho_atual.nome] += 0.5**grau
        if coelho_atual.get_pai():
            self.calc_coeficiente_macho(coelho_inicial, coelho_atual.get_pai(), grau+1)
        if coelho_atual.get_mae():
            self.calc_coeficiente_macho(coelho_inicial, coelho_atual.get_mae(), grau+1)
        for filho in coelho_atual.filhos:
            self.calc_coeficiente_macho(coelho_inicial, filho, grau+1)

    def calcular(self):
        femeas = [coelho.nome for coelho in self.coelhos.values() if coelho.get_sexo() == 'F']
        machos = [coelho.nome for coelho in self.coelhos.values() if coelho.get_sexo() == 'M']
        print(f'Fêmeas: {femeas}')
        print(f'Machos: {machos}')

        # Inicializa o mapa de coeficientes de parentesco
        self.coeficiente_parentesco = defaultdict(dict)
        for macho in machos:
            self.coeficiente_parentesco[macho] = {femea: 0 for femea in femeas}
        # Para cada coelho macho, calcula o coeficiente de parentesco
        for coelho in self.coelhos.values():
            if coelho.get_sexo() == 'M':
                self.visitados = defaultdict(bool)
                print(f'Calculando coeficientes de parentesco para o coelho {coelho.nome}')
                self.calc_coeficiente_macho(coelho)
        
    def get_coeficientes_parentesco(self):
        return self.coeficiente_parentesco

# %% Cálculo dos coeficientes de parentesco
import pandas as pd

cp = CoeficienteParentesco(coelhos)
cp.calcular()

coeficientes_parentesco = cp.get_coeficientes_parentesco()
df = pd.DataFrame(coeficientes_parentesco)
df.index.name = 'coelhos'
df = df.applymap(lambda x: round(x, 4))
df[df<0.0015] = 0
df


#%%
from openpyxl.formatting.rule import ColorScaleRule
EXCEL_SHEET_NAME = 'Coeficientes de parentesco'

with pd.ExcelWriter(EXCEL_FILE_NAME, engine="openpyxl", mode='a') as writer:
    if EXCEL_SHEET_NAME in writer.book.sheetnames:
        idx = writer.book.sheetnames.index(EXCEL_SHEET_NAME)
        writer.book.remove(writer.book.worksheets[idx])

    df.to_excel(writer, sheet_name=EXCEL_SHEET_NAME)

    workbook = writer.book
    worksheet = writer.sheets[EXCEL_SHEET_NAME]

    ultima_celula = worksheet.cell(
        row=df.shape[0]+1, 
        column=df.shape[1]+1
    ).coordinate
    cell_range = f'B2:{ultima_celula}'
    
    worksheet.conditional_formatting.add(cell_range, 
        ColorScaleRule(
            start_type='num', start_value=0, start_color='375623',
            mid_type='num', mid_value=0.25, mid_color='ffc000',
            end_type='num', end_value=0.7, end_color='FF0000'
        )
    )

#%%