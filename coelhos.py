import pandas as pd
from openpyxl import load_workbook
from unicodedata import normalize
from collections import defaultdict
from openpyxl.formatting.rule import ColorScaleRule
# import pygraphviz as PG
import os
import shutil

def get_sheets(filename):
    wb = load_workbook(filename, read_only=True)
    sheets = wb.sheetnames
    return sheets

def get_dataframe(filename, sheet):
    df = pd.read_excel(filename, sheet_name=sheet, dtype=str)

    # Columns to lower case and unicode
    df.columns = [normalize('NFKD', col.lower()).encode('ASCII', 'ignore').decode('ASCII') for col in df.columns]
    
    missing_columns = []
    if 'individuo' not in df.columns: missing_columns.append('individuo')
    if 'pai' not in df.columns: missing_columns.append('pai')
    if 'mae' not in df.columns: missing_columns.append('mae')
    if 'sexo' not in df.columns: missing_columns.append('sexo')
    if len(missing_columns) > 0:
        raise Exception(f'Planilha {sheet} não contém as colunas: {", ".join(missing_columns)}')

    df = df.rename(columns={
        'Indivíduo': 'individuo',
        'Pai': 'pai',
        'Mae': 'mae',
        'SEXO': 'sexo',
    })
    df = df[['individuo', 'pai', 'mae', 'sexo']]

    df.replace('x', 'X', inplace=True)
    df.replace('X', '', inplace=True)
    df['sexo'] = df['sexo'].str.upper()

    df['individuo'] = df['individuo'].apply(lambda x: x.rjust(10, ' ') if x.isnumeric() else x)
    df.sort_values(by=['individuo'], inplace=True)
    df['individuo'] = df['individuo'].str.strip()

    return df

class Coelho:
    def __init__(self, nome, sexo):
        self.nome = nome
        self.sexo = sexo
        self.pai = None
        self.mae = None
        self.filhos = []

    def __repr__(self):
        string = f'Coelho {self.nome} ({self.sexo})\n'
        string+= f'Descendentes: \n{self.get_arvore()}'
        return string

    def add_pai(self, pai):
        if pai.get_sexo() != 'M':
            raise Exception(f'Pai do coelho {self.nome}, coelho {pai.nome}, é fêmea')
        self.pai = pai
        pai.add_filho(self)
    
    def add_mae(self, mae):
        if mae.get_sexo() != 'F':
            raise Exception(f'Mãe do coelho {self.nome}, coelho {mae.nome}, é macho')
        self.mae = mae
        mae.add_filho(self)

    def add_filho(self, filho):
        self.filhos.append(filho)

    def get_arvore(self, nivel=0):
        arvore = f'{"  "*nivel}{self.nome} ({self.sexo})\n'
        for filho in self.filhos:
            arvore += filho.get_arvore(nivel+1)
        return arvore
    def get_pai(self):
        return self.pai
    def get_mae(self):
        return self.mae
    def get_sexo(self):
        return self.sexo

class GrauParentesco:
    def __init__(self, filename, sheet):
        self.filename = filename
        self.sheet_coelhos = sheet
        self.sheet_parentesco = 'grau_parentesco'
        self.sheet_detalhes = 'coeficientes_detalhes'

        # self.grafo_parentesco = PG.AGraph(directed=True, strict=True)
        self.dfCoelhos = get_dataframe(filename, sheet)
        self.coelhos = self.criar_coelhos()
        self.init_coeficiente()
        self.calcular()

    def init_coeficiente(self):
        self.coeficiente_parentesco = defaultdict(dict)
        
        machos = [coelho for coelho in self.coelhos.values() if coelho.get_sexo() == 'M']
        for macho in machos:
            self.init_coeficiente_coelho(macho)
        
    def init_coeficiente_coelho(self, coelho):
        coelhos_opostos = [c.nome for c in self.coelhos.values() if c.get_sexo() != coelho.get_sexo()]
        self.coeficiente_parentesco[coelho.nome] = {coelho_oposto: 0 for coelho_oposto in coelhos_opostos}


    @property
    def get_data_input(self):
        return self.dfCoelhos.to_dict('records')
    
    @property
    def get_data_output(self):
        return self.coeficientes_detalhados
    
    def criar_coelhos(self):
        coelhos = {}
        for index, row in self.dfCoelhos.iterrows():
            nome = row['individuo'].strip()
            sexo = row['sexo'].strip()
            coelhos[nome] = Coelho(nome, sexo)

        for index, row in self.dfCoelhos.iterrows():
            nome = row['individuo'].strip()
            pai = row['pai'].strip()
            mae = row['mae'].strip()
            if pai:
                coelhos[nome].add_pai(coelhos[pai])
                # self.grafo_parentesco.add_edge(pai, nome)
            if mae:
                coelhos[nome].add_mae(coelhos[mae])
                # self.grafo_parentesco.add_edge(mae, nome)
        return coelhos

    def calc_coeficiente_individual(self, coelho_inicial, coelho_atual=None, grau=0, descendente=False):
        if coelho_atual == None: 
            coelho_atual = coelho_inicial
            self._historico = []
            self.visitados = defaultdict(bool)
            self.init_coeficiente_coelho(coelho_inicial)

        self.visitados[coelho_atual.nome] = True
        direcao = '\\'[0] if descendente else '/'
        direcao = '' if coelho_atual == coelho_inicial else direcao
        self._historico.append(direcao + coelho_atual.nome)

        # Se o coelho atual for do sexo oposto ao coelho inicial
        if coelho_atual.get_sexo() != coelho_inicial.get_sexo():
            self.coeficiente_parentesco[coelho_inicial.nome][coelho_atual.nome] += 0.5**grau
            self.coeficientes_detalhados.append({
                'origem': coelho_inicial.nome,
                'destino': coelho_atual.nome,
                'parentesco': ''.join(self._historico.copy()),
                'coeficiente': 0.5**grau
            })
        
        for filho in coelho_atual.filhos:
            if not self.visitados[filho.nome]:
                self.calc_coeficiente_individual(coelho_inicial, filho, grau+1, descendente=True)

        if coelho_atual.get_pai() and not descendente:
            self.calc_coeficiente_individual(coelho_inicial, coelho_atual.get_pai(), grau+1)
        if coelho_atual.get_mae() and not descendente:
            self.calc_coeficiente_individual(coelho_inicial, coelho_atual.get_mae(), grau+1)
        
        self.visitados[coelho_atual.nome] = False
        self._historico.pop()

    def calcular(self):
        self.coeficientes_detalhados = []

        # Para cada coelho macho, calcula o coeficiente de parentesco
        for coelho in self.coelhos.values():
            if coelho.get_sexo() == 'M':
                print(f'Calculando coeficientes de parentesco para o coelho {coelho.nome}')
                self.calc_coeficiente_individual(coelho)

    def get_parentescos(self):
        return self._parentescos
    
    def copy_sheet(self, filename):
        shutil.copy(self.filename, filename)
        self.filename = filename

    def salvar_arvore(self):
        pass
        # self.grafo_parentesco.layout(prog='dot')
        # self.grafo_parentesco.draw('arvore.png')

    def salvar_coeficientes(self):
        # Cria o dataframe
        df = pd.DataFrame(self.coeficiente_parentesco)
        df.fillna(0, inplace=True)
        df.index.name = 'coelhos'
        df = df.applymap(lambda x: round(x, 4))
        df[df<0.0015] = 0

        # Salva o dataframe no arquivo
        with pd.ExcelWriter(self.filename, engine="openpyxl", mode='a') as writer:
            # Remove worksheet se já existir
            if self.sheet_parentesco in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(self.sheet_parentesco)
                writer.book.remove(writer.book.worksheets[idx])

            # Salva o dataframe no arquivo
            df.to_excel(writer, sheet_name=self.sheet_parentesco)

            # Formatação condicional
            workbook = writer.book
            worksheet = writer.sheets[self.sheet_parentesco]

            ultima_celula = worksheet.cell(
                row=df.shape[0]+1, 
                column=df.shape[1]+1
            ).coordinate
            cell_range = f'B2:{ultima_celula}'
            
            worksheet.conditional_formatting.add(cell_range, 
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='375623',
                    mid_type='num', mid_value=0.25, mid_color='ffc000',
                    end_type='num', end_value=0.7, end_color='FF0000'
                )
            )

    def salvar_coeficientes_detalhados(self):
        df = pd.DataFrame(self.coeficientes_detalhados)
        df.sort_values(by=['origem', 'destino'], inplace=True)
        
        # Salva o dataframe no arquivo
        with pd.ExcelWriter(self.filename, engine="openpyxl", mode='a') as writer:
            # Remove worksheet se já existir
            if self.sheet_detalhes in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(self.sheet_detalhes)
                writer.book.remove(writer.book.worksheets[idx])

            # Salva o dataframe no arquivo
            df.to_excel(writer, sheet_name=self.sheet_detalhes, index=False)